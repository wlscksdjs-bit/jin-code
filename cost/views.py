from rest_framework import viewsets, status
from rest_framework.decorators import action, api_view
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import parser_classes
from rest_framework.parsers import MultiPartParser
from django.db.models import Sum, Q
from django.http import HttpResponse
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
import calendar
from .models import CostCategory, Budget, Vendor, Expense
from .serializers import (
    CostCategorySerializer, BudgetSerializer, BudgetCreateSerializer,
    VendorSerializer, ExpenseSerializer, ExpenseCreateSerializer
)


class CostCategoryViewSet(viewsets.ModelViewSet):
    queryset = CostCategory.objects.all()
    serializer_class = CostCategorySerializer
    permission_classes = [IsAuthenticated]


class VendorViewSet(viewsets.ModelViewSet):
    queryset = Vendor.objects.all()
    serializer_class = VendorSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return Vendor.objects.filter(is_active=True)


class BudgetViewSet(viewsets.ModelViewSet):
    queryset = Budget.objects.all()
    serializer_class = BudgetSerializer
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.action == 'create':
            return BudgetCreateSerializer
        return BudgetSerializer

    def get_queryset(self):
        project_id = self.request.query_params.get('project')
        if project_id:
            return Budget.objects.filter(project_id=project_id).select_related('category', 'project')
        return Budget.objects.none()

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=False, methods=['get'])
    def matrix(self, request):
        from projects.models import Project
        from django.db.models import Sum
        
        projects = Project.objects.filter(status__in=['planning', 'in_progress']).order_by('-created_at')[:10]
        categories = CostCategory.objects.filter(is_active=True).order_by('category_type', 'name')
        
        matrix_data = []
        
        for category in categories:
            row = {
                'category_id': category.id,
                'category_name': category.name,
                'category_type': category.category_type,
                'category_type_display': category.get_category_type_display(),
            }
            
            for project in projects:
                budget = Budget.objects.filter(project=project, category=category).first()
                if budget:
                    expenses = budget.expenses.filter(status='approved')
                    used = expenses.aggregate(total=Sum('amount'))['total'] or 0
                    pending = budget.expenses.filter(status='pending').aggregate(total=Sum('amount'))['total'] or 0
                    remaining = float(budget.planned_amount) - float(used) - float(pending)
                    
                    row[f'project_{project.id}'] = {
                        'budget_id': budget.id,
                        'planned': float(budget.planned_amount),
                        'used': float(used),
                        'pending': float(pending),
                        'remaining': remaining,
                        'status': 'overrun' if remaining < 0 else 'warning' if remaining < float(budget.planned_amount) * 0.2 else 'ok'
                    }
                else:
                    row[f'project_{project.id}'] = None
            
            matrix_data.append(row)
        
        project_summaries = []
        for project in projects:
            budgets = Budget.objects.filter(project=project)
            total_planned = sum(float(b.planned_amount) for b in budgets)
            total_used = sum(
                budget.expenses.filter(status='approved').aggregate(total=Sum('amount'))['total'] or 0 
                for budget in budgets
            )
            total_pending = sum(
                budget.expenses.filter(status='pending').aggregate(total=Sum('amount'))['total'] or 0 
                for budget in budgets
            )
            
            project_summaries.append({
                'project_id': project.id,
                'project_name': project.name,
                'client': project.client,
                'status': project.status,
                'total_planned': total_planned,
                'total_used': total_used,
                'total_pending': total_pending,
                'total_remaining': total_planned - total_used - total_pending,
            })
        
        return Response({
            'categories': [{'id': c.id, 'name': c.name, 'type': c.category_type, 'type_display': c.get_category_type_display()} for c in categories],
            'projects': project_summaries,
            'matrix': matrix_data,
        })

    @action(detail=True, methods=['get'])
    def summary(self, request, pk=None):
        budget = self.get_object()
        expenses = budget.expenses.filter(status='approved')
        total_used = expenses.aggregate(total=Sum('amount'))['total'] or 0
        total_pending = budget.expenses.filter(status='pending').aggregate(total=Sum('amount'))['total'] or 0
        
        return Response({
            'planned': str(budget.planned_amount),
            'used': str(total_used),
            'pending': str(total_pending),
            'remaining': str(budget.planned_amount - (total_used + total_pending)),
            'variance': str(budget.planned_amount - total_used),
            'variance_status': 'savings' if budget.planned_amount > total_used else 'overrun'
        })

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        from projects.models import Project
        
        projects = Project.objects.filter(status__in=['planning', 'in_progress', 'completed']).order_by('-created_at')[:20]
        
        wb = Workbook()
        ws = wb.active
        ws.title = "프로젝트 원가 관리"
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        headers = ['프로젝트', '고객사', '상태', '비용항목', '계획금액', '사용금액', '대기금액', '잔액', '상태']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        
        row = 2
        for project in projects:
            budgets = Budget.objects.filter(project=project).select_related('category')
            if budgets:
                for budget in budgets:
                    expenses = budget.expenses.filter(status='approved')
                    used = float(expenses.aggregate(total=Sum('amount'))['total'] or 0)
                    pending = float(budget.expenses.filter(status='pending').aggregate(total=Sum('amount'))['total'] or 0)
                    remaining = float(budget.planned_amount) - used - pending
                    
                    status_text = '초과' if remaining < 0 else '주의' if remaining < float(budget.planned_amount) * 0.2 else '정상'
                    
                    ws.cell(row=row, column=1, value=project.name).border = thin_border
                    ws.cell(row=row, column=2, value=project.client).border = thin_border
                    ws.cell(row=row, column=3, value=project.get_status_display()).border = thin_border
                    ws.cell(row=row, column=4, value=budget.category.name).border = thin_border
                    ws.cell(row=row, column=5, value=float(budget.planned_amount)).border = thin_border
                    ws.cell(row=row, column=6, value=used).border = thin_border
                    ws.cell(row=row, column=7, value=pending).border = thin_border
                    ws.cell(row=row, column=8, value=remaining).border = thin_border
                    ws.cell(row=row, column=9, value=status_text).border = thin_border
                    row += 1
            else:
                ws.cell(row=row, column=1, value=project.name).border = thin_border
                ws.cell(row=row, column=2, value=project.client).border = thin_border
                ws.cell(row=row, column=3, value=project.get_status_display()).border = thin_border
                row += 1
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 15
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['D'].width = 15
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="project_cost_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        return response

    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser])
    def import_excel(self, request):
        from projects.models import Project
        from django.contrib.auth import get_user_model
        User = get_user_model()
        
        file = request.FILES.get('file')
        if not file:
            return Response({'error': '파일이 없습니다.'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            wb = Workbook(BytesIO(file.read()))
            ws = wb.active
            
            created_projects = []
            errors = []
            
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                if not row[0]:
                    continue
                
                project_name = row[0]
                client = row[1] or ''
                status_val = row[2] or 'waiting'
                
                project, created = Project.objects.get_or_create(
                    name=project_name,
                    defaults={
                        'client': client,
                        'status': status_val if status_val in ['waiting', 'planning', 'in_progress', 'completed', 'cancelled'] else 'waiting',
                        'created_by': request.user,
                    }
                )
                
                if created:
                    created_projects.append(project_name)
                else:
                    errors.append(f'{row_num}행: "{project_name}" 프로젝트가 이미 존재합니다.')
            
            return Response({
                'message': f'{len(created_projects)}개 프로젝트 생성됨',
                'created': created_projects,
                'errors': errors
            })
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['get'])
    def monthly_report(self, request, pk=None):
        project = Project.objects.get(pk=pk)
        budgets = Budget.objects.filter(project=project).select_related('category')
        
        year = int(request.query_params.get('year', datetime.now().year))
        month = int(request.query_params.get('month', datetime.now().month))
        
        _, last_day = calendar.monthrange(year, month)
        start_date = datetime(year, month, 1).date()
        end_date = datetime(year, month, last_day).date()
        
        report_data = []
        total_planned = 0
        total_used_prev = 0
        total_used_current = 0
        total_pending = 0
        
        for budget in budgets:
            prev_expenses = budget.expenses.filter(
                status='approved',
                expense_date__lt=start_date
            )
            current_expenses = budget.expenses.filter(
                status='approved',
                expense_date__gte=start_date,
                expense_date__lte=end_date
            )
            pending_expenses = budget.expenses.filter(status='pending')
            
            used_prev = float(prev_expenses.aggregate(total=Sum('amount'))['total'] or 0)
            used_current = float(current_expenses.aggregate(total=Sum('amount'))['total'] or 0)
            pending = float(pending_expenses.aggregate(total=Sum('amount'))['total'] or 0)
            planned = float(budget.planned_amount)
            
            total_planned += planned
            total_used_prev += used_prev
            total_used_current += used_current
            total_pending += pending
            
            remaining = planned - used_prev - used_current - pending
            usage_rate = ((used_prev + used_current) / planned * 100) if planned > 0 else 0
            
            report_data.append({
                'category_id': budget.category.id,
                'category_name': budget.category.name,
                'category_type': budget.category.category_type,
                'planned': planned,
                'used_previous': used_prev,
                'used_current': used_current,
                'pending': pending,
                'remaining': remaining,
                'usage_rate': round(usage_rate, 1),
                'status': 'overrun' if remaining < 0 else 'warning' if usage_rate > 80 else 'ok'
            })
        
        total_remaining = total_planned - total_used_prev - total_used_current - total_pending
        total_usage_rate = ((total_used_prev + total_used_current) / total_planned * 100) if total_planned > 0 else 0
        
        return Response({
            'project': {
                'id': project.id,
                'name': project.name,
                'client': project.client,
                'status': project.status,
            },
            'year': year,
            'month': month,
            'items': report_data,
            'summary': {
                'total_planned': total_planned,
                'total_used_previous': total_used_prev,
                'total_used_current': total_used_current,
                'total_pending': total_pending,
                'total_remaining': total_remaining,
                'total_usage_rate': round(total_usage_rate, 1),
                'status': 'overrun' if total_remaining < 0 else 'warning' if total_usage_rate > 80 else 'ok'
            }
        })

    @action(detail=True, methods=['post'])
    def create_approval_from_report(self, request, pk=None):
        from projects.models import Project
        from approvals.models import Approval, ApprovalType
        
        project = Project.objects.get(pk=pk)
        year = int(request.data.get('year', datetime.now().year))
        month = int(request.data.get('month', datetime.now().month))
        
        budgets = Budget.objects.filter(project=project).select_related('category')
        
        total_used = 0
        total_planned = 0
        for budget in budgets:
            total_planned += float(budget.planned_amount)
            total_used += float(budget.expenses.filter(status='approved').aggregate(total=Sum('amount'))['total'] or 0)
        
        usage_rate = (total_used / total_planned * 100) if total_planned > 0 else 0
        variance = total_planned - total_used
        status_text = '초과' if variance < 0 else '정상' if variance >= total_planned * 0.2 else '주의'
        
        approval_type = ApprovalType.objects.filter(code='cost_report').first()
        if not approval_type:
            approval_type = ApprovalType.objects.first()
        
        month_name = f'{year}년 {month}월'
        
        content = f"""[{month_name} 원가 분석 보고서]

프로젝트: {project.name}
고객사: {project.client}

【요약】
- 계획금액: {total_planned:,.0f}원
- 사용금액: {total_used:,.0f}원
- 사용률: {usage_rate:.1f}%
- 잔액: {variance:,.0f}원
- 상태: {status_text}

【상세 내역】
"""
        for budget in budgets:
            cat_used = float(budget.expenses.filter(status='approved').aggregate(total=Sum('amount'))['total'] or 0)
            cat_planned = float(budget.planned_amount)
            cat_rate = (cat_used / cat_planned * 100) if cat_planned > 0 else 0
            content += f"- {budget.category.name}: {cat_used:,.0f} / {cat_planned:,.0f} ({cat_rate:.1f}%)\n"
        
        approval = Approval.objects.create(
            project=project,
            approval_type=approval_type,
            title=f'[{month_name}] {project.name} 원가 분석 보고서',
            content=content,
            amount=str(total_used),
            status='pending',
            created_by=request.user
        )
        
        return Response({
            'message': '결재 상신 완료',
            'approval_id': approval.id,
            'status': approval.status
        })


class ExpenseViewSet(viewsets.ModelViewSet):
    queryset = Expense.objects.all()
    serializer_class = ExpenseSerializer
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.action == 'create':
            return ExpenseCreateSerializer
        return ExpenseSerializer

    def get_queryset(self):
        budget_id = self.request.query_params.get('budget')
        project_id = self.request.query_params.get('project')
        
        queryset = Expense.objects.select_related('budget__category', 'budget__project', 'vendor', 'submitted_by', 'approved_by')
        
        if budget_id:
            queryset = queryset.filter(budget_id=budget_id)
        elif project_id:
            queryset = queryset.filter(budget__project_id=project_id)
        
        return queryset

    def perform_create(self, serializer):
        serializer.save(submitted_by=self.request.user)

    @action(detail=True, methods=['post'])
    def submit(self, request, pk=None):
        expense = self.get_object()
        expense.status = 'submitted'
        expense.save()
        return Response(ExpenseSerializer(expense).data)

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        expense = self.get_object()
        if not request.user.can_approve:
            return Response({'error': '승인 권한이 없습니다.'}, status=status.HTTP_403_FORBIDDEN)
        expense.status = 'approved'
        expense.approved_by = request.user
        expense.save()
        return Response(ExpenseSerializer(expense).data)

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        expense = self.get_object()
        if not request.user.can_approve:
            return Response({'error': '승인 권한이 없습니다.'}, status=status.HTTP_403_FORBIDDEN)
        expense.status = 'rejected'
        expense.approved_by = request.user
        expense.save()
        return Response(ExpenseSerializer(expense).data)

    @action(detail=False, methods=['get'], url_path='project/(?P<project_id>[^/.]+)')
    def by_project(self, request, project_id=None):
        expenses = Expense.objects.filter(budget__project_id=project_id).select_related(
            'budget__category', 'budget__project', 'vendor', 'submitted_by', 'approved_by'
        )
        return Response(ExpenseSerializer(expenses, many=True).data)